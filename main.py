from distutils import command
from html.parser import HTMLParser

import os, sys
from unicodedata import category
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import date
from datetime import datetime
import datetime
import time

import tkinter
from tkinter import *
import tkinter.ttk as ttk
from tkinter import messagebox

import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule

import platform
import pyshorteners


nameCount = 1
priceCount = 1
index = 1

nowDate = date.today().isoformat()
now=datetime.datetime.now()


printTitles = ''
printPrice = ''

titleList = []
shopList = []
trendList = []
pageList = []

keywordList = ['노트북', '반려동물', '자동차용품', '청소용품']

global comboList
wb = Workbook()


url = ''
cateNum = ''

options = webdriver.ChromeOptions()
#options.add_experimental_option('excludeSwitches', ['enable-logging'])
#options.add_argument("headless")
#driver = webdriver.Chrome(executable_path='./chromedriver', options=options)

    

def make_clickable(val):
    return '<a href="{}">{}</a>'.format(val,val)

def activeLabel(text):
    pProgress['text'] = text
        
def scanPlatform():
    scanOS = platform.system()
    if scanOS == 'Windows':
        pass
    elif scanOS == 'Darwin':
        pass


def filtering_string(value, index):
  
    
    stringValue = value.replace(str(index)+'위','').replace('유지','').replace('상품','').replace('펼치기','').replace('상품접기','').replace('상승','').replace('하락','').replace('접기','')
   
    return stringValue
   
    

def doScrollDown(whileSeconds, driver):
    
    start = datetime.datetime.now()
    end = start + datetime.timedelta(seconds=whileSeconds)
    
    element = driver.find_element_by_tag_name('html')
    
    for i in range(800):
        element.send_keys(Keys.ARROW_DOWN)
    
    #while True:
        #driver.execute_script('window.scrollTo(0, document.body.scrollHeight);')
        #time.sleep(2)
        #if datetime.datetime.now() > end:
            #break
        
        

def comboFunc(readVal):
    
    if readVal == '노트북':
        url = 'https://search.shopping.naver.com/best/category/keyword?categoryCategoryId=50000151&categoryChildCategoryId=&categoryDemo=A00&categoryMidCategoryId=50000151&categoryRootCategoryId=50000003&chartRank=1&period=P7D'
        cateNum = '50000151'
        return url, cateNum
    elif readVal == '반려동물':
        url = 'https://search.shopping.naver.com/best/category/keyword?categoryCategoryId=50000155&categoryChildCategoryId=&categoryDemo=A00&categoryMidCategoryId=50000155&categoryRootCategoryId=50000008&chartRank=1&period=P7D'
        cateNum = '50000155'
        return url, cateNum
    elif readVal == '자동차용품':
        url = 'https://search.shopping.naver.com/best/category/keyword?categoryCategoryId=50000055&categoryChildCategoryId=&categoryDemo=A00&categoryMidCategoryId=50000055&categoryRootCategoryId=50000008&chartRank=1&period=P7D'
        cateNum = '50000055'
        return url, cateNum
    elif readVal == '청소용품':
        url = 'https://search.shopping.naver.com/best/category/keyword?categoryCategoryId=50000077&categoryChildCategoryId=&categoryDemo=A00&categoryMidCategoryId=50000077&categoryRootCategoryId=50000008&chartRank=1&period=P7D'
        cateNum = '50000077'
    else: pass
    
    
    





def crawlKeyword():
    
    activeLabel('에러발생!! 나중에 다시 시도해주세요 :)')
    time.sleep(1)    
    nameCount = 1
    priceCount = 1
    trendList = []
    cateNum = ''
    driver = webdriver.Chrome(executable_path='./chromedriver', options=options)

    readCombo = comboList.get()
    print(readCombo)
    
    url, cateNum = comboFunc(readCombo)
    
    # driver.implicitly_wait(3)
    # url에 접근한다.
    driver.get(url)

    #주간 버튼 클릭 (일간일때는 요놈만 주석처리하거나 나중에 선택지를 주는 방법으로 선택가능)
    #driver.find_element_by_xpath("//*[@id='__next']/div/div[2]/div[2]/div/div[2]/div[2]/a[2]").click()


    #doScrollDown(2,driver) #2초간 스크롤 넉넉하게 2초 동안 스크롤함

    req = driver.page_source #스크롤 한 후 페이지 소스 불러오기
    soup = BeautifulSoup(req, 'html.parser')

    trends = soup.select('a.chartList_btn_keyword__1F7BO')
    
    #lowPrices = soup.select('div.imageProduct_price__3vXjm') #최저가
    


    print('트렌드 키워드 순위')

    for trend in trends:
        
        
        printTrend = trend.get_text()
    
 
        filterTrend = filtering_string(printTrend,nameCount)
        #print(nameCount)
        print(filterTrend)
        
       
    
        nameCount = nameCount+1
    
        trendList.append(filterTrend)
    
    
    print(trendList)
    
    activeLabel('키워드 수집 성공')

    
    #dest_filename = str(nowDate)
    
    
    #for i in range(len(trendList)):
        
        #wb.create_sheet(trendList[i])
        
        
    #wb.save(dest_filename+'.xlsx')
    #wb.close()
    
    for i in range(len(trendList)):
        driver.get('https://search.shopping.naver.com/search/all?query='+trendList[i]+'&catId='+cateNum)
        titleList=[]
        shopList=[]
        pageList=[]
        for pageNum in range(1, 11):
            doScrollDown(2, driver)
        
            req = driver.page_source
            soup = BeautifulSoup(req, 'html.parser')
        
            titles = soup.select('div.basicList_title__3P9Q7') #제품 이름 출력
            shopLink = soup.select('div.basicList_mall_title__3MWFY') #쇼핑 링크

            num = 1
            for title in titles:
                printTitle = title.get_text()
                
                
                titleList.append(str(pageNum)+'---'+str(num)+'---'+printTitle)
                num = num + 1
            
            print()    
            print(titleList) 
     
            
            for shopName in shopLink:
                printShop = shopName.get_text()
            
                shopList.append(printShop)
            
            
            driver.find_element_by_css_selector('#__next > div > div.style_container__1YjHN > div.style_inner__18zZX > div.style_content_wrap__1PzEo > div.style_content__2T20F > div.pagination_pagination__6AcG4 > a.pagination_next__1ITTf').click()
            
        list_of_tuples = list(zip(titleList,shopList))
        df = pd.DataFrame(list_of_tuples, columns=['페이지---순위---제품명','쇼핑몰이름'])
        df.index = df.index+1
        
        if not os.path.exists('./'+readCombo+nowDate+'.xlsx'):
            with pd.ExcelWriter('./'+readCombo+nowDate+'.xlsx', mode='w', engine='openpyxl') as writer:
                df.to_excel(writer, header=True, index= True, sheet_name=trendList[i],index_label='순위')
        else:
            with pd.ExcelWriter('./'+readCombo+nowDate+'.xlsx', mode='a', engine='openpyxl') as writer:
                df.to_excel(writer, header=True, index= True, sheet_name=trendList[i],index_label='순위')
        
    
    driver.quit()


def searchKeyword():
    
    readCombo = comboList.get()
    
    
    url, cateNum = comboFunc(readCombo)
    
    searchRes = textBox.get()
    print(searchRes)
    if searchRes == '':
        messagebox.showwarning('키워드가 없습니다','키워드를 입력해주세요!')
    else:
        driver = webdriver.Chrome(executable_path='./chromedriver', options=options)

        driver.get('https://search.shopping.naver.com/search/all?query='+searchRes+'&catId='+cateNum)
        titleList=[]
        shopList=[]
        pageList=[]
        for pageNum in range(1, 11):
            doScrollDown(2, driver)
        
            req = driver.page_source
            soup = BeautifulSoup(req, 'html.parser')
        
            titles = soup.select('div.basicList_title__3P9Q7') #제품 이름 출력
            shopLink = soup.select('div.basicList_mall_title__3MWFY') #쇼핑 링크

            num = 1
            for title in titles:
                printTitle = title.get_text()
                    
                
                titleList.append(str(pageNum)+'---'+str(num)+'---'+printTitle)
                num = num + 1
            
            print()    
            print(titleList) 
     
            
            for shopName in shopLink:
                printShop = shopName.get_text()
        
                shopList.append(printShop)
            
            
            driver.find_element_by_css_selector('#__next > div > div.style_container__1YjHN > div.style_inner__18zZX > div.style_content_wrap__1PzEo > div.style_content__2T20F > div.pagination_pagination__6AcG4 > a.pagination_next__1ITTf').click()
            
        list_of_tuples = list(zip(titleList,shopList))
        df = pd.DataFrame(list_of_tuples, columns=['페이지---순위---제품명','쇼핑몰이름'])
        df.index = df.index+1
        
        if not os.path.exists('./'+searchRes+nowDate+'.xlsx'):
            with pd.ExcelWriter('./'+searchRes+nowDate+'.xlsx', mode='w', engine='openpyxl') as writer:
                df.to_excel(writer, header=True, index= True, sheet_name=searchRes,index_label='순위')
        else:
            with pd.ExcelWriter('./'+searchRes+'-'+nowDate+'.xlsx', mode='a', engine='openpyxl') as writer:
                df.to_excel(writer, header=True, index= True, sheet_name=searchRes,index_label='순위')
        
    
        driver.quit()
    activeLabel('키워드 수집 성공')
    

def relationKeyword():
    
    readCombo = comboList.get()
    
    
    url, cateNum = comboFunc(readCombo)
    
    searchRes = textBox.get()
    print(searchRes)
    if searchRes == '':
        messagebox.showwarning('키워드가 없습니다','키워드를 입력해주세요!')
    else:
        driver = webdriver.Chrome(executable_path='./chromedriver', options=options)

        driver.get('https://search.shopping.naver.com/search/all?query='+searchRes+'&catId='+cateNum)
        titleList=[]
        naverSearchList=[]

        
        
        req = driver.page_source
        soup = BeautifulSoup(req, 'html.parser')
        
        titles = soup.select('div.relatedTags_relation_srh__1CleC ul li') 

        num = 1
        for title in titles:
            printTitle = title.get_text()
                    
                
            titleList.append(printTitle)
            num = num + 1
            
        print()    
        print(titleList) 
        
        driver.get('https://search.naver.com/search.naver?sm=tab_hty.top&where=nexearch&query='+searchRes)
        req = driver.page_source
        soup = BeautifulSoup(req, 'html.parser')
        
        releateKey = soup.select('a.keyword')
        
        for releate in releateKey:
            printReleate = releate.get_text()
            filterKeyword = printReleate.replace(' ','')
            naverSearchList.append(filterKeyword)
        
        print(naverSearchList)
        
         

        dic = {'네이버쇼핑 연관':titleList,'네이버검색 연관':naverSearchList}
        df = pd.DataFrame.from_dict(dic, orient='index')
    
        #df.index = df.index+1
        df = df.transpose()
        
        if not os.path.exists('./'+'연관)'+searchRes+nowDate+'.xlsx'):
            with pd.ExcelWriter('./'+'연관)'+searchRes+nowDate+'.xlsx', mode='w', engine='openpyxl') as writer:
                df.to_excel(writer, header=True, index= True, sheet_name=searchRes,index_label='번호')
        else:
            with pd.ExcelWriter('./'+'연관)'+searchRes+nowDate+'.xlsx', mode='a', engine='openpyxl') as writer:
                df.to_excel(writer, header=True, index= True, sheet_name=searchRes,index_label='번호')
        
    
        driver.quit()
    activeLabel('키워드 수집 성공')
    


window = tkinter.Tk()
    
window.title('트렌드 키워드 순위 검색기')
window.geometry('300x300+300+200')
window.resizable(False,False)
pTitle = Label(window, text = '트렌드 키워드 순위 검색기')

textBox = Entry(window, width= 20)
    
pProgress = Label(window, text= "버전 1.3")
    
initBtn = Button(window, text='트렌드 키워드 전체 데이터 수집', command=crawlKeyword)
searchBtn = Button(window, text='사용자 지정\n키워드 수집', command=searchKeyword, width=10, height=3)
releateBtn = Button(window, text='연관 키워드\n수집', command=relationKeyword, width = 10, height=3)
comboList = ttk.Combobox(window, height=5, values=keywordList, state="readonly")
comboList.current(0)
    
pTitle.pack()
comboList.pack()
initBtn.pack(side='top')

textBox.pack(padx=2, pady=10)
releateBtn.pack(padx=2)
searchBtn.pack(padx=2)

pProgress.pack(side='top', pady=3)





    
window.mainloop()



        
    

    
    
    